import pandas as pd
import numpy as np
from utils.date_utils import get_days_in_current_month, FIRST_DAY_OF_THE_MONTH_IS_SATURDAY, get_holiday_array
from utils.model_utils import get_taken_tasks_per_operator
from utils.stats_utils import get_operator_filtered_tasks
from utils.operator_utils import dont_want_task, want_task, is_operator_strong_no_task

from datetime import timedelta, date
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, PatternFill, Fill, Font, Border, Alignment, Side
from openpyxl.styles.borders import BORDER_THICK, BORDER_MEDIUMDASHED

import datetime

from output.statistics import add_statistics


# dataframe for each operator as row


def create_operators_dataframe(operators):
    df = pd.DataFrame(columns=operators.name)
    for date in get_days_in_current_month():
        df = df.append({"Date": date}, ignore_index=True)
    return df


# dataframe for each task as a row


def create_task_dataframe(DB_path):
    data_df = pd.read_excel(DB_path, sheet_name='Tasks')
    df = pd.DataFrame(columns=data_df.name)
    for date in get_days_in_current_month():
        df = df.append({"Date": date}, ignore_index=True)
    return df, data_df


# converts binary matrix to dfs by operator and by task
def convert_to_readable_df(shifts_model, tasks, operators, DB_path):
    operators_df = create_operators_dataframe(operators)
    tasks_df_multiple = create_task_dataframe(DB_path)
    tasks_df = tasks_df_multiple[0]
    tasks_df_colors = tasks_df_multiple[1]
    df = pd.DataFrame()
    by_date_dict = create_shift_dict(shifts_model, tasks, operators, DB_path)
    # needs to get the shifts model
    # parent key is date key is operator and value is task
    for parent_key, parent_value in by_date_dict.items():
        if type(parent_value) is dict:
            for key, value in parent_value.items():
                task_start_time = parent_key
                operator_name = key
                task_name = value
                df.at[operator_name, task_start_time] = remove_cell_sign(task_name)
                # if this is sofash write the next day too
                cell_y = task_name.split("&")[1]
                if ("_" in task_name):
                    cell_y = task_name.split("_")[0].split("&")[1]
                cell_y = int(cell_y)
                nextDay = tasks['start_time'][cell_y] + timedelta(days=1)
                next_day_start_time = str(nextDay).split(' ')[0]

                if 'Sofash' in task_name or 'holiday' in task_name:
                    df.at[operator_name, next_day_start_time] = remove_cell_sign(task_name)
                    nextDay = nextDay + timedelta(days=1)
                    next_day_start_time = str(nextDay).split(' ')[0]
                    df.at[operator_name, next_day_start_time] = 'MALAM'
                elif 'night' in task_name:
                    if (nextDay.weekday() != 4):
                        df.at[operator_name, next_day_start_time] = 'MALAM'
    for index, row in operators.iterrows():
        df.at[row['name'], 'pazam'] = row['MAX']
        # df.at[operator['name']]
    colors_dict = create_colors_dict(tasks_df_colors)

    df = df.sort_index(1)
    add_statistics(shifts_model, df, operators, tasks)
    df = df.sort_values(by='pazam', ascending=False)
    df = df.drop('pazam', axis='columns')
    return df, colors_dict


def remove_cell_sign(task_name):
    if "_" not in task_name:
        return task_name.split("&")[0]
    split_tasks = task_name.split("_")
    new_name = ""
    for i in range(len(split_tasks)):
        if i < len(split_tasks) - 1:
            new_name += split_tasks[i].split("&")[0] + "_"
        else:
            new_name += split_tasks[i].split("&")[0]
    return new_name


def create_shift_dict(shifts_model, tasks, operators, DB_path):
    operators_df = create_operators_dataframe(operators)
    tasks_df_multiple = create_task_dataframe(DB_path)
    tasks_df = tasks_df_multiple[0]
    tasks_df_colors = tasks_df_multiple[1]

    df = pd.DataFrame()
    by_date_dict = {}
    # needs to get the shifts model

    for i, v in enumerate([v for v in shifts_model.vars if v.name[0] == 'x']):
        if v.x >= 0.99:
            cell = v.name[2:]
            cell = cell[:-1]
            cell_x = int(cell.split(",")[0])
            cell_y = int(cell.split(",")[1])
            operator_name = operators['name'][cell_x]
            task_name = tasks['name'][cell_y]
            task_start_time = str(tasks['start_time'][cell_y]).split(' ')[0]
            if by_date_dict.get(task_start_time) == None:
                by_date_dict[task_start_time] = {}
            if by_date_dict[task_start_time].get(operator_name) == None:
                by_date_dict[task_start_time][operator_name] = task_name + "&" + str(cell_y)
            else:
                by_date_dict[task_start_time][operator_name] = by_date_dict[task_start_time][
                                                                   operator_name] + "_" + task_name + "&" + str(cell_y)
            # if by_date_dict.get(task_start_time)
    return by_date_dict


def create_colors_dict(color_df):
    colors_dict_from_df = color_df.to_dict()
    colors_dict = {}
    for i in range(len(colors_dict_from_df["name"])):
        colors_dict[colors_dict_from_df["name"][i]
        ] = colors_dict_from_df["color"][i][1:]
    return colors_dict


def color_cells(file_path, shifts_model, color_dict, operators, tasks):
    new_book = Workbook()
    new_sheet = new_book.active
    book = load_workbook(file_path)
    ws = book.worksheets[0]

    color_tasks(color_dict, new_sheet, ws)

    color_special_dates(new_sheet, ws, shifts_model, operators, tasks, color='00FF0000',
                        special_date_name="Not evening", filter_method=dont_want_task, border_type=BORDER_MEDIUMDASHED)
    color_special_dates(new_sheet, ws, shifts_model, operators, tasks, color='00FF0000',
                        filter_method=is_operator_strong_no_task, special_date_name="Not task")
    color_special_dates(new_sheet, ws, shifts_model, operators, tasks, color='0000FF00', filter_method=want_task,
                        special_date_name="Preferred days")

    return new_book.save('./output/Butzi.xlsx')


def color_special_dates(new_sheet, ws, shifts_model, operators, tasks, color, special_date_name, filter_method,
                        border_type=BORDER_THICK):
    oper_names = [col[0].value for col in ws.columns]
    taken_tasks_per_operator = get_taken_tasks_per_operator(shifts_model)
    for oper_index, operator in operators.iterrows():
        oper_id = oper_names.index(operator['name'])
        tasks_to_color_df = get_operator_filtered_tasks(oper_index, operator, taken_tasks_per_operator, tasks,
                                                        filter_method=filter_method)
        for special_date in str(operator[special_date_name]).split(','):
            if special_date == 'nan':
                continue
            border = Border(
                left=Side(border_style=border_type, color=color),
                right=Side(border_style=border_type, color=color),
                top=Side(border_style=border_type, color=color),
                bottom=Side(border_style=border_type, color=color)
            )

            cell_to_color = ws.cell(row=int(special_date) + 1 - FIRST_DAY_OF_THE_MONTH_IS_SATURDAY,
                                    column=oper_id + 1)
            new_cell = new_sheet.cell(row=cell_to_color.row, column=cell_to_color.column, value=cell_to_color.value)
            new_cell.border = border
            if new_cell.value:
                for task_name in new_cell.value.split('_'):
                    if task_name in list(tasks_to_color_df['name']):
                        new_cell.font = Font(color=color, bold=True)
                        continue


def color_tasks(color_dict, new_sheet, ws):
    for col in ws:
        for cell in col:
            try:
                new_cell = new_sheet.cell(
                    row=cell.row, column=cell.column, value=cell.value)
                # color by task
                if cell.value in color_dict:
                    new_cell.fill = PatternFill(
                        start_color=color_dict[cell.value], end_color=color_dict[cell.value], fill_type="solid")
                # add weekend colors
                elif cell.value == 'MALAM':
                    new_cell.fill = PatternFill(
                        start_color='e5e5e5', end_color='f5f5f5', fill_type='solid')
                elif cell.row > 1 and (
                        datetime.datetime.strptime(
                            ws.cell(cell.row, 1, ).value, '%Y-%m-%d').weekday() in [4, 5] or datetime.datetime.strptime(
                    ws.cell(cell.row, 1, ).value, '%Y-%m-%d').day in get_holiday_array()):
                    new_cell.fill = PatternFill(
                        start_color="e4e4e4", end_color="e4e4e4", fill_type="solid")
                    new_cell.font = Font(bold=True)
                # bold for titles
                else:
                    new_cell.font = Font(bold=True)
            except:
                pass
