from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from utils.date_utils import get_days_in_current_month, us_day_to_il_day, FIRST_DAY_OF_THE_MONTH_IS_SATURDAY, \
    get_holiday_array

from constants import WEEKDAYS_LIST


def add_tasks_sheets(file_path, tasks_df, operators_df):
    add_tahak_sheet(file_path, tasks_df, operators_df)
    add_equipment_sheet(file_path, tasks_df, operators_df)
    add_bakut_sheet(file_path, tasks_df, operators_df)
    add_production_sheet(file_path, tasks_df, operators_df)


def add_tahak_sheet(file_path, tasks_df, operators_df):
    output_workbook = load_workbook(file_path)
    output_workbook.create_sheet('cleaning the tahak')
    tahak_worksheet = output_workbook.get_sheet_by_name('cleaning the tahak')
    tasks_types = {
        'date': [],
        'day': [],
        'cleaning the tahak': ['CLEANING THE TAHAK']
    }
    create_sheet_data(file_path, operators_df, output_workbook, tahak_worksheet, tasks_df, tasks_types)


def add_equipment_sheet(file_path, tasks_df, operators_df):
    output_workbook = load_workbook(file_path)
    output_workbook.create_sheet('equipment')
    equipment_worksheet = output_workbook.get_sheet_by_name('equipment')
    tasks_types = {
        'date': [],
        'day': [],
        'equipment operator': ['equip operator'],
        'equipment day': ['equipment day', ('Sofash Equip day', True), ('holiday Equip day', True)],
        'equipment night': ['equipment night', 'Hamishi - equipment night', ('Sofash Equip night', True),
                            ('holiday Equip night', True)],
        'WV operator': ['wv equipment']
    }

    create_sheet_data(file_path, operators_df, output_workbook, equipment_worksheet, tasks_df, tasks_types)


def add_bakut_sheet(file_path, tasks_df, operators_df):
    output_workbook = load_workbook(file_path)
    output_workbook.create_sheet('bakut')
    bakut_worksheet = output_workbook.get_sheet_by_name('bakut')

    tasks_types = {
        'date': [],
        'day': [],
        'bakut operator': ['bakut operator'],
        'bakut day 1': ['bakut day 1', ('Sofash bakut day 1', True), ('holiday bakut day 1', True)],
        'bakut day 2': ['bakut day 2', ('Sofash bakut day 2', True), ('holiday bakut day 2', True)],
        'bakut night 1': ['bakut night 1', 'Hamishi - bakut night 1', ('Sofash bakut night 1', True),
                          ('holiday bakut night 1', True)],
        'bakut night 2': ['bakut night 2', 'Hamishi - bakut night 2', ('Sofash bakut night 2', True),
                          ('holiday bakut night 2', True)]
    }

    create_sheet_data(file_path, operators_df, output_workbook, bakut_worksheet, tasks_df, tasks_types)


def add_production_sheet(file_path, tasks_df, operators_df):
    output_workbook = load_workbook(file_path)
    output_workbook.create_sheet('production')
    production_worksheet = output_workbook.get_sheet_by_name('production')

    tasks_types = {
        'date': [],
        'day': [],
        'eo production': ['eo production', ('Sofash production', True), ('holiday production', True)],
        'sar production': ['sar production'],
        'eo production operator': ['eo production operator'],
        'sar production operator': ['sar production operator']
    }

    create_sheet_data(file_path, operators_df, output_workbook, production_worksheet, tasks_df, tasks_types)


def add_sheet_titles(worksheet_to_edit, titles: list):
    for title_index, title in enumerate(titles):
        worksheet_to_edit.cell(row=1, column=title_index + 1).value = title


def add_dates_column(worksheet_to_edit, row_length):
    for date in get_days_in_current_month():
        day_name = WEEKDAYS_LIST[us_day_to_il_day(date.weekday())]
        add_weekend_color(worksheet_to_edit, date, row_length=row_length)
        worksheet_to_edit.cell(row=date.day + 1 - FIRST_DAY_OF_THE_MONTH_IS_SATURDAY, column=1).value = str(date)
        worksheet_to_edit.cell(row=date.day + 1 - FIRST_DAY_OF_THE_MONTH_IS_SATURDAY, column=2).value = day_name


def add_weekend_color(worksheet_to_edit, date, row_length):
    if us_day_to_il_day(date.weekday()) in [5, 6] or date.day in get_holiday_array():
        for column_index in range(1, row_length + 1):
            worksheet_to_edit.cell(row=date.day + 1 - FIRST_DAY_OF_THE_MONTH_IS_SATURDAY, column=column_index).fill = \
                PatternFill(start_color="e4e4e4", end_color="e4e4e4", fill_type="solid")


def assign_operators_to_tasks_cells(worksheet_to_edit, tasks_df, operators_df, column_to_edit, is_weekend=False):
    for _, task in tasks_df.iterrows():
        worksheet_to_edit.cell(row=task['start_time'].day + 1 - FIRST_DAY_OF_THE_MONTH_IS_SATURDAY,
                               column=column_to_edit).value = \
            operators_df.iloc[task['assignee']]['name']
        if is_weekend:
            worksheet_to_edit.cell(row=task['start_time'].day + 2 - FIRST_DAY_OF_THE_MONTH_IS_SATURDAY,
                                   column=column_to_edit).value = \
                operators_df.iloc[task['assignee']]['name']


def blacken_empty_cells(worksheet_to_edit):
    for col in worksheet_to_edit:
        for cell in col:
            if not cell.value:
                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")


def create_sheet_data(file_path, operators_df, output_workbook, worksheet_to_edit, tasks_df, tasks_types):
    add_sheet_titles(worksheet_to_edit, list(tasks_types.keys()))
    add_dates_column(worksheet_to_edit, row_length=len(tasks_types.keys()))
    for index, (key, task_type) in enumerate(tasks_types.items()):
        for task in task_type:
            is_weekand = True if len(task) > 1 and task[1] == True else False
            if is_weekand:
                task = task[0]

            assign_operators_to_tasks_cells(worksheet_to_edit, tasks_df.loc[tasks_df.name == task], operators_df,
                                            column_to_edit=index + 1, is_weekend=is_weekand)
    blacken_empty_cells(worksheet_to_edit)
    output_workbook.save(file_path)
